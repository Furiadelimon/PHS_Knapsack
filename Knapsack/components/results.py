"""
Módulo que contiene los componentes para visualizar los resultados del problema de la mochila.
"""

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from static.styles import backpack_html

def display_results(results, max_weight, solve_time=None):
    """
    Muestra los resultados de la optimización de la mochila con un diseño profesional y sobrio.
    
    Args:
        results (dict): Diccionario con los resultados del solucionador
        max_weight (int): Capacidad máxima de la mochila
        solve_time (float, optional): Tiempo de ejecución del solucionador
    """
    # Extraer los datos de los resultados
    items_selected = results['selected_items']
    total_benefit = results['total_benefit']
    total_weight = results['total_weight']
    weights = results['weights']
    benefits = results['benefits']
    
    # Crear DataFrame para exportación
    data = []
    selected_items_binary = results.get('selected_items_binary', [0]*len(benefits))
    for i in range(len(benefits)):
        data.append({
            'Ítem': f'Ítem {i+1}',
            'Seleccionado': '✅' if selected_items_binary[i] else '❌',
            'Seleccionado_valor': 1 if selected_items_binary[i] else 0,
            'Beneficio': benefits[i],
            'Peso': weights[i]
        })
    export_df = pd.DataFrame(data)
    
    # Guardar el DataFrame en session_state para exportación
    st.session_state.export_df = export_df
    st.session_state.results_summary = {
        'total_benefit': total_benefit,
        'total_weight': total_weight,
        'max_weight': max_weight,
        'efficiency': round(total_benefit / total_weight, 2) if total_weight > 0 else 0,
        'items_selected': sum(selected_items_binary),
        'total_items': len(benefits),
        'algorithm_used': st.session_state.algorithm_used,
        'solve_time': solve_time
    }
    
    # Configuración visual de la sección de resultados
    st.markdown("""
    <div style="background-color: #f9f9f9; 
                padding: 1rem; 
                border-radius: 5px; 
                border: 1px solid #ddd;
                margin: 1rem 0;">
        <div style="font-size: 0.9rem; color: #2c3e50; font-weight: 600; margin-bottom: 0.8rem; display: flex; align-items: center;">
            <span style="margin-right: 6px;">✅</span> Resultados de la Optimización
        </div>
    """, unsafe_allow_html=True)
    
    # Crear una distribución de columnas más compacta y elegante
    col1, col2, col3 = st.columns([2, 1, 2])
    
    # Columna 1: Tabla de resultados
    with col1:
        # Crear tabla de resultados detallados
        data = []
        for i in range(len(benefits)):
            data.append({
                'Ítem': f'Ítem {i+1}',
                'Seleccionado': '✅' if selected_items_binary[i] else '❌',
                'Beneficio': benefits[i],
                'Peso': weights[i]
            })
        df = pd.DataFrame(data)
        # Agregar una fila de totales
        totals = pd.DataFrame([{
            'Ítem': 'TOTAL',
            'Seleccionado': f"{sum(selected_items_binary)}/{len(selected_items_binary)}",
            'Beneficio': total_benefit,
            'Peso': total_weight
        }])
        # Concatenar con la fila de totales
        df = pd.concat([df, totals], ignore_index=True)
        
        # Título de la sección
        st.markdown("""
        <div style="font-size: 0.85rem; color: #2c3e50; font-weight: 600; margin-bottom: 0.5rem;">
            📋 Detalle de la Solución
        </div>
        """, unsafe_allow_html=True)
        
        # Mostrar la tabla con estilo
        st.dataframe(
            df,
            column_config={
                "Ítem": st.column_config.TextColumn("Ítem"),
                "Seleccionado": st.column_config.TextColumn("Seleccionado"),
                "Beneficio": st.column_config.NumberColumn("Beneficio", format="%d"),
                "Peso": st.column_config.NumberColumn("Peso", format="%d")
            },
            hide_index=True,
            use_container_width=True,
        )
    
    # Columna 2: Mostrar métricas clave y mochila visual
    with col2:
        # Título de la sección
        st.markdown("""
        <div style="font-size: 0.85rem; color: #2c3e50; font-weight: 600; margin-bottom: 0.5rem; text-align: center;">
            📊 Resumen
        </div>
        """, unsafe_allow_html=True)
        
        # Calcular la eficiencia (beneficio/peso)
        efficiency = round(total_benefit / total_weight, 2) if total_weight > 0 else 0
        
        # Métricas de resumen con estilo más compacto y profesional
        st.markdown(f"""
        <div style="background-color: #ffffff; padding: 0.7rem; border-radius: 4px; border: 1px solid #eee; margin-bottom: 0.6rem; text-align: center;">
            <div style="font-size: 0.7rem; color: #7f8c8d; margin-bottom: 0.2rem;">Beneficio Total</div>
            <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{total_benefit}</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="background-color: #ffffff; padding: 0.7rem; border-radius: 4px; border: 1px solid #eee; margin-bottom: 0.6rem; text-align: center;">
            <div style="font-size: 0.7rem; color: #7f8c8d; margin-bottom: 0.2rem;">Peso Utilizado</div>
            <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{total_weight} <span style="font-size: 0.8rem; color: #7f8c8d;">/ {max_weight}</span></div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="background-color: #ffffff; padding: 0.7rem; border-radius: 4px; border: 1px solid #eee; margin-bottom: 0.6rem; text-align: center;">
            <div style="font-size: 0.7rem; color: #7f8c8d; margin-bottom: 0.2rem;">Eficiencia</div>
            <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{efficiency} <span style="font-size: 0.7rem; color: #7f8c8d;">beneficio/peso</span></div>
        </div>
        """, unsafe_allow_html=True)

        if solve_time is not None:
            st.markdown(f"""
            <div style="background-color: #ffffff; padding: 0.7rem; border-radius: 4px; border: 1px solid #eee; text-align: center;">
                <div style="font-size: 0.7rem; color: #7f8c8d; margin-bottom: 0.2rem;">⏱️ Tiempo de Ejecución</div>
                <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{solve_time:.4f} <span style="font-size: 0.7rem; color: #7f8c8d;">segundos</span></div>
            </div>
            """, unsafe_allow_html=True)

    # Columna 3: Gráficos
    with col3:
        # Título de la sección
        st.markdown("""
        <div style="font-size: 0.85rem; color: #2c3e50; font-weight: 600; margin-bottom: 0.5rem;">
            📈 Visualización Gráfica
        </div>
        """, unsafe_allow_html=True)
        
        # Crear las pestañas para los diferentes gráficos con estilo más sobrio
        tab1, tab2 = st.tabs(["Comparativa", "Distribución"])
        
        with tab1:
            # Gráfico de comparativa
            fig, ax = plt.subplots(figsize=(8, 5))
            
            # Preparar datos para el gráfico
            df_plot = pd.DataFrame({
                'Ítem': [f'Ítem {i+1}' for i in range(len(benefits))],
                'Beneficio': benefits,
                'Peso': weights,
                'Seleccionado': ['Seleccionado' if s else 'No seleccionado' for s in results.get('selected_items_binary', [0]*len(benefits))]
            })
            
            # Gráfico con Seaborn y colores más sobrios
            sns.set_theme(style="whitegrid")
            bar_plot = sns.barplot(
                data=df_plot,
                x='Ítem',
                y='Beneficio',
                hue='Seleccionado',
                palette={'Seleccionado': '#27ae60', 'No seleccionado': '#f5f5f5'},
                ax=ax
            )
            
            # Ajustar el gráfico
            plt.ylabel('Beneficio')
            plt.xlabel('Ítem')
            plt.title('Comparativa de Beneficios por Ítem', fontsize=10, color='#2c3e50')
            plt.xticks(rotation=45)
            plt.legend(title='Estado', loc='upper right', frameon=True, framealpha=0.7)
            plt.tight_layout()
            
            # Mostrar el gráfico
            st.pyplot(fig)
        
        with tab2:
            # Gráfico de distribución
            fig, ax = plt.subplots(figsize=(8, 5))
            
            # Crear datos para gráfico circular de pesos usados vs capacidad restante
            weight_used = total_weight
            weight_remaining = max_weight - total_weight
            
            # Datos para el gráfico circular con colores más sobrios
            labels = ['Peso Utilizado', 'Peso Disponible']
            sizes = [weight_used, weight_remaining]
            colors = ['#2c3e50', '#ecf0f1']
            explode = (0.1, 0)  # Destacar el peso utilizado
            
            # Crear el gráfico
            plt.pie(
                sizes, 
                explode=explode, 
                labels=labels, 
                colors=colors, 
                autopct='%1.1f%%',
                shadow=False, 
                startangle=90,
                textprops={'color': '#2c3e50', 'fontsize': 9}
            )
            plt.axis('equal')
            plt.title('Distribución de Capacidad de la Mochila', fontsize=10, color='#2c3e50')
            
            # Mostrar el gráfico
            st.pyplot(fig)
    
    # Mensaje de conclusión más elegante
    st.markdown(f"""
    <div style="background-color: #edf8ff; padding: 0.8rem; border-radius: 4px; border-left: 3px solid #2c3e50; margin-top: 1rem;">
        <div style="font-size: 0.85rem; color: #2c3e50;">
            <span style="font-weight: 600;">Conclusión:</span> Se han seleccionado <span style="font-weight: 600;">{sum(items_selected)}</span> ítems de un total de <span style="font-weight: 600;">{len(items_selected)}</span>, 
            obteniendo un beneficio total de <span style="font-weight: 600;">{total_benefit}</span> unidades y utilizando <span style="font-weight: 600;">{total_weight}</span> de <span style="font-weight: 600;">{max_weight}</span> unidades de peso disponible.
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Cerrar el div principal
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Acciones adicionales con estilo más sobrio
    col_actions1, col_actions2, col_actions3 = st.columns(3)
    
    with col_actions1:
        if st.button("🔄 Resolver Otro Problema", use_container_width=True):
            # Reiniciar estado para resolver otro problema
            st.session_state.results = None
            st.session_state.selected_algorithm = "ortools"
            st.rerun()
    
    with col_actions2:
        if st.button("📊 Generar Informe PDF", use_container_width=True):
            generate_pdf_report()
    
    with col_actions3:
        if st.button("📥 Exportar a Excel", use_container_width=True):
            export_to_excel()

def generate_pdf_report():
    """
    Genera un informe PDF con los resultados de la optimización.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
        import matplotlib.pyplot as plt
        import base64
        import datetime
        
        # Verificar que existan resultados para generar el informe
        if 'export_df' not in st.session_state or 'results_summary' not in st.session_state:
            st.error("No hay resultados disponibles para generar el informe.")
            return
        
        # Crear un buffer para el PDF
        buffer = io.BytesIO()
        
        # Configurar el documento
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Center', alignment=1, fontSize=12))
        # Usar un nombre diferente para evitar el conflicto con el estilo predefinido 'Title'
        styles.add(ParagraphStyle(name='ReportTitle', alignment=1, fontSize=16, fontName='Helvetica-Bold'))
        styles.add(ParagraphStyle(name='Subtitle', alignment=1, fontSize=14, fontName='Helvetica-Bold'))
        
        # Elementos del PDF
        elements = []
        
        # Título - usar el nuevo nombre de estilo 'ReportTitle'
        elements.append(Paragraph("Informe de Optimización de la Mochila", styles['ReportTitle']))
        elements.append(Spacer(1, 0.25*inch))
        
        # Fecha actual
        now = datetime.datetime.now()
        date_str = now.strftime("%d/%m/%Y %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {date_str}", styles['Center']))
        elements.append(Spacer(1, 0.25*inch))
        
        # Resumen
        summary = st.session_state.results_summary
        elements.append(Paragraph("Resumen de Resultados", styles['Subtitle']))
        elements.append(Spacer(1, 0.1*inch))
        
        summary_data = [
            ["Algoritmo utilizado:", summary['algorithm_used']],
            ["Beneficio total:", str(summary['total_benefit'])],
            ["Peso utilizado:", f"{summary['total_weight']} / {summary['max_weight']}"],
            ["Eficiencia (beneficio/peso):", str(summary['efficiency'])],
            ["Ítems seleccionados:", f"{summary['items_selected']} / {summary['total_items']}"],
            ["Tiempo de ejecución:", f"{summary['solve_time']:.4f} segundos" if summary['solve_time'] else "N/A"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.25*inch))
        
        # Tabla detallada
        elements.append(Paragraph("Detalle de Ítems", styles['Subtitle']))
        elements.append(Spacer(1, 0.1*inch))
        
        df = st.session_state.export_df
        detail_data = [["Ítem", "Seleccionado", "Beneficio", "Peso"]]
        for _, row in df.iterrows():
            detail_data.append([
                row['Ítem'],
                "Sí" if row['Seleccionado_valor'] == 1 else "No",
                row['Beneficio'],
                row['Peso']
            ])
        
        # Agregar fila de totales
        detail_data.append([
            "TOTAL",
            f"{summary['items_selected']} / {summary['total_items']}",
            summary['total_benefit'],
            summary['total_weight']
        ])
        
        detail_table = Table(detail_data)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(detail_table)
        elements.append(Spacer(1, 0.25*inch))
        
        # Crear gráficos para el PDF
        fig, ax = plt.subplots(figsize=(7, 4))
        df_selected = df[df['Seleccionado_valor'] == 1]
        df_not_selected = df[df['Seleccionado_valor'] == 0]
        
        # Gráfico de barras para los beneficios de los ítems seleccionados vs no seleccionados
        ax.bar(df_selected['Ítem'], df_selected['Beneficio'], color='#27ae60', label='Seleccionados')
        ax.bar(df_not_selected['Ítem'], df_not_selected['Beneficio'], color='#e74c3c', label='No seleccionados')
        ax.set_xlabel('Ítem')
        ax.set_ylabel('Beneficio')
        ax.set_title('Beneficios por Ítem')
        ax.legend()
        plt.tight_layout()
        
        # Guardar gráfico en buffer
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png')
        img_buffer.seek(0)
        
        # Añadir gráfico al PDF
        elements.append(Paragraph("Visualización Gráfica", styles['Subtitle']))
        elements.append(Spacer(1, 0.1*inch))
        
        img = Image(img_buffer, width=6*inch, height=3*inch)
        elements.append(img)
        
        # Pie de página
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("PHS Knapsack - Informe generado automáticamente", styles['Center']))
        
        # Construir el PDF
        doc.build(elements)
        
        # Preparar para descarga
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        
        # Crear el enlace de descarga
        file_name = f"Informe_Mochila_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
        href = f'<a href="data:application/pdf;base64,{b64}" download="{file_name}">📥 Haga clic aquí para descargar el informe PDF</a>'
        
        # Mostrar mensaje y enlace de descarga
        st.success("Informe PDF generado correctamente.")
        st.markdown(href, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"Error al generar el informe PDF: {str(e)}")
        # Sugerir instalación si es un error de importación
        if "No module named" in str(e):
            st.info("Asegúrese de tener instalado reportlab: `pip install reportlab`")

def export_to_excel():
    """
    Exporta los resultados a un archivo Excel.
    """
    try:
        import io
        import base64
        import datetime
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, Series
        
        # Verificar que existan resultados para exportar
        if 'export_df' not in st.session_state or 'results_summary' not in st.session_state:
            st.error("No hay resultados disponibles para exportar.")
            return
        
        # Crear un buffer para el archivo Excel
        buffer = io.BytesIO()
        
        # Obtener los datos
        df = st.session_state.export_df
        summary = st.session_state.results_summary
        
        # Crear un libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resultados"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        selected_fill = PatternFill(start_color="edf7ed", end_color="edf7ed", fill_type="solid")
        total_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
        center_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Título
        sheet.merge_cells('A1:F1')
        sheet['A1'] = "Resultados de Optimización de la Mochila"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha
        now = datetime.datetime.now()
        date_str = now.strftime("%d/%m/%Y %H:%M:%S")
        sheet.merge_cells('A2:F2')
        sheet['A2'] = f"Generado el: {date_str}"
        sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        sheet.merge_cells('A4:F4')
        sheet['A4'] = "Resumen de Resultados"
        sheet['A4'].font = Font(bold=True, size=12)
        sheet['A4'].alignment = Alignment(horizontal='center')
        
        # Datos de resumen
        summary_data = [
            ["Algoritmo utilizado:", summary['algorithm_used']],
            ["Beneficio total:", summary['total_benefit']],
            ["Peso utilizado:", f"{summary['total_weight']} / {summary['max_weight']}"],
            ["Eficiencia (beneficio/peso):", summary['efficiency']],
            ["Ítems seleccionados:", f"{summary['items_selected']} / {summary['total_items']}"],
            ["Tiempo de ejecución:", f"{summary['solve_time']:.4f} segundos" if summary['solve_time'] else "N/A"]
        ]
        
        # Escribir resumen
        for i, row in enumerate(summary_data, 5):
            sheet.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
            sheet.cell(row=i, column=2, value=row[1])
        
        # Título tabla de detalles
        sheet.merge_cells('A12:F12')
        sheet['A12'] = "Detalle de Ítems"
        sheet['A12'].font = Font(bold=True, size=12)
        sheet['A12'].alignment = Alignment(horizontal='center')
        
        # Encabezados de tabla
        headers = ["Ítem", "Seleccionado", "Beneficio", "Peso"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=13, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Datos de detalle
        for i, (_, row) in enumerate(df.iterrows(), 14):
            # Ítem
            cell = sheet.cell(row=i, column=1, value=row['Ítem'])
            cell.alignment = center_alignment
            cell.border = border
            
            # Seleccionado
            selected_value = "Sí" if row['Seleccionado_valor'] == 1 else "No"
            cell = sheet.cell(row=i, column=2, value=selected_value)
            cell.alignment = center_alignment
            cell.border = border
            if row['Seleccionado_valor'] == 1:
                cell.fill = selected_fill
            
            # Beneficio
            cell = sheet.cell(row=i, column=3, value=row['Beneficio'])
            cell.alignment = center_alignment
            cell.border = border
            
            # Peso
            cell = sheet.cell(row=i, column=4, value=row['Peso'])
            cell.alignment = center_alignment
            cell.border = border
        
        # Fila de totales
        total_row = i + 1
        
        cell = sheet.cell(row=total_row, column=1, value="TOTAL")
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = total_fill
        
        cell = sheet.cell(row=total_row, column=2, value=f"{summary['items_selected']} / {summary['total_items']}")
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = total_fill
        
        cell = sheet.cell(row=total_row, column=3, value=summary['total_benefit'])
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = total_fill
        
        cell = sheet.cell(row=total_row, column=4, value=summary['total_weight'])
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = total_fill
        
        # Crear una hoja para gráficos
        chart_sheet = workbook.create_sheet(title="Gráficos")
        
        # Preparar datos para el gráfico
        chart_sheet['A1'] = "Ítem"
        chart_sheet['B1'] = "Beneficio"
        chart_sheet['C1'] = "Seleccionado"
        
        for i, (_, row) in enumerate(df.iterrows(), 2):
            chart_sheet.cell(row=i, column=1, value=row['Ítem'])
            chart_sheet.cell(row=i, column=2, value=row['Beneficio'])
            chart_sheet.cell(row=i, column=3, value="Seleccionado" if row['Seleccionado_valor'] == 1 else "No seleccionado")
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.title = "Beneficio por Ítem"
        chart.x_axis.title = "Ítem"
        chart.y_axis.title = "Beneficio"
        
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=i, max_col=2)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=i)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        chart_sheet.add_chart(chart, "E5")
        
        # Guardar el libro
        workbook.save(buffer)
        
        # Preparar para descarga
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        
        # Crear el enlace de descarga
        file_name = f"Resultados_Mochila_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">📥 Haga clic aquí para descargar el archivo Excel</a>'
        
        # Mostrar mensaje y enlace de descarga
        st.success("Archivo Excel generado correctamente.")
        st.markdown(href, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"Error al exportar a Excel: {str(e)}")
        # Sugerir instalación si es un error de importación
        if "No module named" in str(e):
            st.info("Asegúrese de tener instalado openpyxl: `pip install openpyxl`")

def display_comparison_table(results_dict, max_weight):
    """
    Muestra una tabla comparativa entre diferentes algoritmos para el problema de la mochila.
    
    Args:
        results_dict (dict): Diccionario donde cada clave es el nombre del algoritmo
                            y cada valor es el diccionario de resultados de ese algoritmo
        max_weight (int): Capacidad máxima de la mochila
    """
    import pandas as pd
    import streamlit as st
    import matplotlib.pyplot as plt
    import seaborn as sns
    import numpy as np
    
    # Extraer métricas para cada algoritmo
    metrics = []
    for algo_name, results in results_dict.items():
        metrics.append({
            'Algoritmo': algo_name,
            'Beneficio Total': results['total_benefit'],
            'Peso Utilizado': results['total_weight'],
            'Eficiencia (Beneficio/Peso)': round(results['total_benefit'] / results['total_weight'], 2) if results['total_weight'] > 0 else 0,
            'Tiempo de Ejecución (ms)': round(results['execution_time'], 3),
            'Ítems Seleccionados': sum(results.get('selected_items_binary', results['selected_items'])),
            'Óptimo': 'Sí' if 'Óptimo' in algo_name else 'No'
        })
    
    # Crear DataFrame
    df_metrics = pd.DataFrame(metrics)
    
    # Configurar el estilo de la tabla con un diseño más moderno y compacto
    st.markdown("""
    <div style="background-color: #f9f9f9; border: 1px solid #eaeaea; padding: 0.7rem; 
                border-radius: 5px; margin: 0.5rem 0; box-shadow: 0 1px 3px rgba(0,0,0,0.05);">
        <div style="font-size: 0.85rem; font-weight: 600; color: #2c3e50; margin-bottom: 0.5rem; 
                    display: flex; align-items: center;">
            <span style="margin-right: 6px;">📊</span> Comparación de Algoritmos
        </div>
    """, unsafe_allow_html=True)
    
    # Mostrar la tabla con formato mejorado
    st.dataframe(
        df_metrics,
        column_config={
            "Algoritmo": st.column_config.TextColumn("Algoritmo"),
            "Beneficio Total": st.column_config.NumberColumn("Beneficio", format="%d"),
            "Peso Utilizado": st.column_config.NumberColumn("Peso", format="%d"),
            "Eficiencia (Beneficio/Peso)": st.column_config.NumberColumn("Eficiencia", format="%.2f"),
            "Tiempo de Ejecución (ms)": st.column_config.NumberColumn("Tiempo", format="%.3f ms"),
            "Ítems Seleccionados": st.column_config.NumberColumn("Ítems", format="%d"),
            "Óptimo": st.column_config.TextColumn("Óptimo")
        },
        hide_index=True,
        use_container_width=True,
    )
    
    # Visualizaciones comparativas con diseño mejorado
    col1, col2 = st.columns([5, 3])
    
    with col1:
        # Gráfico de barras comparativo - TAMAÑO REDUCIDO Y MEJOR DISEÑO
        fig, ax = plt.subplots(figsize=(6, 3))  # Tamaño reducido de la figura
        
        # Preparar datos para visualización
        algorithms = [m['Algoritmo'] for m in metrics]
        benefits = [m['Beneficio Total'] for m in metrics]
        times = [m['Tiempo de Ejecución (ms)'] for m in metrics]
        
        # Configurar el estilo del gráfico
        sns.set_theme(style="whitegrid", palette="pastel", font_scale=0.8)
        x = np.arange(len(algorithms))
        width = 0.35  # Ancho de barra reducido
        
        # Gráfico de barras para beneficio con colores más elegantes
        bar1 = ax.bar(x - width/2, benefits, width, label='Beneficio', color='#3498db', alpha=0.9, 
                      edgecolor='white', linewidth=0.7)
        
        # Crear un segundo eje Y para los tiempos
        ax2 = ax.twinx()
        bar2 = ax2.bar(x + width/2, times, width, label='Tiempo (ms)', color='#e74c3c', alpha=0.7,
                       edgecolor='white', linewidth=0.7)
        
        # Añadir valores encima de las barras
        for i, v in enumerate(benefits):
            ax.text(i - width/2, v + max(benefits)*0.02, str(v), ha='center', va='bottom', 
                    fontsize=7, color='#3498db', fontweight='bold')
        
        for i, v in enumerate(times):
            ax2.text(i + width/2, v + max(times)*0.02, f"{v:.1f}", ha='center', va='bottom', 
                     fontsize=7, color='#e74c3c', fontweight='bold')
        
        # Añadir etiquetas y leyenda con tamaño reducido
        ax.set_xlabel('Algoritmo', fontsize=8, color='#555')
        ax.set_ylabel('Beneficio', fontsize=8, color='#3498db')
        ax2.set_ylabel('Tiempo (ms)', fontsize=8, color='#e74c3c')
        ax.set_title('Comparativa: Beneficio vs Tiempo', fontsize=9, color='#2c3e50', pad=8)
        ax.set_xticks(x)
        ax.set_xticklabels(algorithms, rotation=35, ha='right', fontsize=7)
        ax.tick_params(axis='y', labelsize=7, colors='#3498db')
        ax2.tick_params(axis='y', labelsize=7, colors='#e74c3c')
        
        # Combinar leyendas de ambos ejes con diseño mejorado
        lines1, labels1 = ax.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=7, 
                 frameon=True, framealpha=0.7, edgecolor='#ddd')
        
        # Mejorar el grid para que sea más sutil
        ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
        
        plt.tight_layout()
        st.pyplot(fig)
    
    with col2:
        # Gráfico de eficiencia comparativa
        fig2, ax = plt.subplots(figsize=(4, 3))
        
        # Extraer datos de eficiencia
        efficiency = [m['Eficiencia (Beneficio/Peso)'] for m in metrics]
        
        # Gráfico de barras horizontal para eficiencia
        bars = ax.barh(algorithms, efficiency, color='#2ecc71', alpha=0.8, 
                      edgecolor='white', linewidth=0.7)
        
        # Añadir valores al final de las barras
        for i, v in enumerate(efficiency):
            ax.text(v + max(efficiency)*0.02, i, f"{v:.2f}", va='center', fontsize=7, 
                   color='#2c3e50', fontweight='bold')
        
        # Configuración del gráfico
        ax.set_xlabel('Eficiencia (Beneficio/Peso)', fontsize=8, color='#555')
        ax.set_title('Eficiencia por Algoritmo', fontsize=9, color='#2c3e50', pad=8)
        ax.tick_params(axis='y', labelsize=7)
        ax.tick_params(axis='x', labelsize=7)
        ax.invert_yaxis()  # Invertir para que coincida con el orden del gráfico anterior
        ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
        
        plt.tight_layout()
        st.pyplot(fig2)
    
    # Añadir breve análisis de los resultados
    st.markdown("""
    <div style="background-color: #edf8ff; padding: 0.6rem; border-radius: 4px; 
                border-left: 3px solid #2c3e50; margin-top: 0.4rem; 
                font-size: 0.78rem; color: #2c3e50;">
        <span style="font-weight: 600;">Análisis Comparativo:</span>
        <ul style="margin: 0.2rem 0 0 0; padding-left: 1.2rem;">
            <li>Los algoritmos óptimos (Programación Dinámica y Branch & Bound) obtienen el mismo beneficio.</li>
            <li>Greedy es significativamente más rápido pero puede dar soluciones subóptimas.</li>
            <li>La eficiencia (beneficio/peso) permite evaluar el aprovechamiento de la capacidad disponible.</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    # Cerrar el div principal
    st.markdown("</div>", unsafe_allow_html=True)

def display_complexity_growth(max_n=30, brute_force_threshold=20, current_n=None):
    """
    Muestra una gráfica comparativa del crecimiento de la complejidad de fuerza bruta (2^n)
    frente a algoritmos eficientes (n^2) para el problema de la mochila.
    Args:
        max_n (int): Valor máximo de n a mostrar en la gráfica
        brute_force_threshold (int): Valor de n a partir del cual fuerza bruta es impracticable
        current_n (int, optional): Valor actual de n seleccionado por el usuario
    """
    import streamlit as st
    import numpy as np
    import matplotlib.pyplot as plt

    n_values = np.arange(1, max_n + 1)
    brute_force = 2 ** n_values
    efficient = n_values ** 2

    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.plot(n_values, brute_force, label='Fuerza Bruta (2ⁿ)', color='#e74c3c', linewidth=2)
    ax.plot(n_values, efficient, label='Algoritmo Eficiente (n²)', color='#3498db', linewidth=2)
    ax.set_yscale('log')
    ax.set_xlabel('n (número de ítems)', fontsize=9)
    ax.set_ylabel('Operaciones (escala log)', fontsize=9)
    ax.set_title('Crecimiento de la Complejidad según n', fontsize=11, color='#2c3e50')
    ax.grid(True, which='both', linestyle='--', linewidth=0.5, alpha=0.5)
    ax.legend(fontsize=8, loc='upper left')

    # Línea vertical para el n actual
    if current_n is not None:
        ax.axvline(current_n, color='#f39c12', linestyle=':', linewidth=2)
        ax.text(current_n+0.2, brute_force[current_n-1], f"n = {current_n}", color='#f39c12', fontsize=8, va='bottom')

    # Sombrear zona donde fuerza bruta es impracticable
    if brute_force_threshold < max_n:
        ax.axvspan(brute_force_threshold, max_n, color='#f8d7da', alpha=0.25)
        ax.text(brute_force_threshold+0.5, brute_force[-1]/3, 'Zona impracticable para fuerza bruta',
                color='#c0392b', fontsize=8, va='top', ha='left', alpha=0.8)

    plt.tight_layout()
    st.pyplot(fig)
    st.markdown("""
    <div style='font-size:0.82rem; color:#2c3e50; background:#f9f9f9; border-radius:4px; border:1px solid #eee; padding:0.7rem; margin-top:0.5rem;'>
        <b>Nota:</b> La complejidad de fuerza bruta crece exponencialmente (2ⁿ), por lo que solo es viable para valores pequeños de n.<br>
        Los algoritmos eficientes (como programación dinámica o OR-Tools) crecen polinómicamente y permiten resolver instancias mucho mayores.
    </div>
    """, unsafe_allow_html=True)